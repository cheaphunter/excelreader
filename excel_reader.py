import os
import fnmatch
import traceback
from openpyxl import load_workbook

class bcolors:
    GREEN = '\033[93m'
    ENDC = '\033[0m'

def search(filename, keyword):
    print("Searching {}".format(filename))
    wb = load_workbook(filename)
    # get all sheets
    for s in wb.get_sheet_names():
        sh = wb[s]
        # get all columns
        for col in range(1, sh.max_column+1):
            # get all rows
            for row_index in range(1, sh.max_row+1):
                if sh.cell(row=row_index, column=col).value == keyword:
                    print(bcolors.GREEN + "[*] Found in {}".format(filename) + bcolors.ENDC)
    print("Finished {}".format(filename))

def get_all_target(path_target):
    file_list = []
    for dir, dn, f in os.walk(path_target):
        for ext in ['xls', 'xlsx']:
            temp_list = fnmatch.filter(os.listdir(dir), '*.'+ext)
            for i in range(len(temp_list)):
                if len(temp_list[i])>0:
                    temp_list[i] = dir+'\\'+temp_list[i]
                    file_list.append(temp_list[i])
    return(file_list)

if __name__ in '__main__':
    # modify two params here and you are ready to go.
    keyword = ''
    path_target = 'd:\\path' # absolute directory
    
    files = get_all_target(path_target)
    for file in files:
        search(file, keyword)
